#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ЭДО ЛДПР - ПОЛНАЯ РАБОЧАЯ ВЕРСИЯ ДЛЯ RENDER
Исправлено: создание распоряжений (отдельная страница)
Все функции работают
"""

import os
import sys
import json
import sqlite3
import secrets
import uuid
import io
import time
from datetime import datetime, timedelta
from functools import wraps

from flask import (
    Flask, render_template_string, request, redirect, url_for,
    session, flash, jsonify, send_file, g
)
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename

# ============================================================
# КОНФИГУРАЦИЯ
# ============================================================

app_flask = Flask(__name__)
app_flask.secret_key = secrets.token_hex(32)
DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'edo_ldpr.db')
app_flask.config['DATABASE'] = DB_PATH
app_flask.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=8)
app_flask.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024

STATIC_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static')
UPLOAD_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads')
ALLOWED_EXTENSIONS = {'pdf', 'doc', 'docx', 'xls', 'xlsx', 'jpg', 'png', 'txt', 'zip'}

os.makedirs(STATIC_DIR, exist_ok=True)
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# ============================================================
# ДЕКОРАТОРЫ
# ============================================================

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            flash('Пожалуйста, войдите в систему', 'warning')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated

def role_required(*roles):
    def decorator(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            if 'user_role' not in session:
                return redirect(url_for('login'))
            if session['user_role'] not in roles:
                flash('Недостаточно прав', 'danger')
                return redirect(url_for('dashboard'))
            return f(*args, **kwargs)
        return decorated
    return decorator

@app_flask.context_processor
def inject_user():
    if 'user_id' in session:
        user = get_db().execute("SELECT * FROM users WHERE uid = ?", (session['user_id'],)).fetchone()
        if user:
            return {'current_user': dict(user)}
    return {'current_user': None}

# ============================================================
# БАЗА ДАННЫХ
# ============================================================

def get_db():
    if 'db' not in g:
        g.db = sqlite3.connect(DB_PATH)
        g.db.row_factory = sqlite3.Row
        g.db.execute("PRAGMA journal_mode=WAL")
        g.db.execute("PRAGMA foreign_keys=ON")
    return g.db

@app_flask.teardown_appcontext
def close_db(error):
    db = g.pop('db', None)
    if db is not None:
        db.close()

def init_db():
    db = get_db()
    db.executescript('''
        CREATE TABLE IF NOT EXISTS users (
            uid TEXT PRIMARY KEY,
            full_name TEXT NOT NULL,
            email TEXT UNIQUE NOT NULL,
            username TEXT UNIQUE,
            password TEXT NOT NULL,
            role TEXT DEFAULT 'executor',
            department_id TEXT,
            avatar_url TEXT,
            is_active INTEGER DEFAULT 1,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS departments (
            id TEXT PRIMARY KEY,
            name TEXT NOT NULL,
            description TEXT,
            head_id TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS orders (
            id TEXT PRIMARY KEY,
            title TEXT NOT NULL,
            content TEXT,
            priority TEXT DEFAULT 'Нормальный',
            status TEXT DEFAULT 'Черновик',
            created_by TEXT NOT NULL,
            creator_name TEXT,
            assigned_department_id TEXT,
            assigned_executor_id TEXT,
            deadline TEXT,
            result TEXT,
            result_file TEXT,
            order_file TEXT,
            rejection_reason TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            updated_at TEXT DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS order_history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            order_id TEXT NOT NULL,
            action TEXT NOT NULL,
            user_name TEXT,
            user_role TEXT,
            details TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS notifications (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id TEXT NOT NULL,
            order_id TEXT,
            message TEXT NOT NULL,
            is_read INTEGER DEFAULT 0,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS settings (
            key TEXT PRIMARY KEY,
            value TEXT
        );
        CREATE TABLE IF NOT EXISTS files (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            order_id TEXT NOT NULL,
            filename TEXT NOT NULL,
            filepath TEXT NOT NULL,
            filetype TEXT,
            filesize INTEGER,
            uploaded_by TEXT,
            uploaded_at TEXT DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS backup_history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            backup_file TEXT NOT NULL,
            backup_size INTEGER,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        );
    ''')
    db.commit()

def seed_database():
    db = get_db()
    user_count = db.execute("SELECT COUNT(*) FROM users").fetchone()[0]
    if user_count > 0:
        return
    
    departments = [
        ('dept-1', 'Центральный аппарат', 'Высшее руководство', None),
        ('dept-2', 'Юридический отдел', 'Правовое обеспечение', None),
        ('dept-3', 'Организационный отдел', 'Организационная работа', None),
        ('dept-4', 'Информационный отдел', 'ИТ и информационная работа', None),
        ('dept-5', 'Отдел регионального развития', 'Работа с регионами', None),
    ]
    for d in departments:
        db.execute("INSERT OR IGNORE INTO departments (id, name, description, head_id) VALUES (?, ?, ?, ?)", d)
    db.commit()
    
    users_data = [
        ('u-admin', 'Администратор Системы', 'admin@ldpr.ru', 'admin', 'admin123', 'admin', None),
        ('u-sec', 'Главный Секретарь', 'sec@ldpr.ru', 'secretary', 'sec123', 'secretary', None),
        ('u-head-central', 'Руководитель ЦА', 'headca@ldpr.ru', 'head_central', 'head123', 'head_central', 'dept-1'),
        ('u-head-dept', 'Начальник Юридического Отдела', 'headlaw@ldpr.ru', 'head_department', 'head123', 'head_department', 'dept-2'),
        ('u-ast', 'Помощник Депутата', 'ast@ldpr.ru', 'assistant', 'ast123', 'assistant', None),
        ('u-exec', 'Рядовой Исполнитель', 'exec@ldpr.ru', 'executor', 'exec123', 'executor', 'dept-2'),
        ('u-exec2', 'Специалист ИТ', 'it@ldpr.ru', 'executor2', 'exec123', 'executor', 'dept-4'),
    ]
    
    for uid, full_name, email, username, plain_pwd, role, dept_id in users_data:
        hashed = generate_password_hash(plain_pwd)
        try:
            db.execute(
                "INSERT INTO users (uid, full_name, email, username, password, role, department_id) VALUES (?, ?, ?, ?, ?, ?, ?)",
                (uid, full_name, email, username, hashed, role, dept_id)
            )
        except sqlite3.IntegrityError:
            db.execute(
                "UPDATE users SET password = ?, role = ?, department_id = ? WHERE username = ?",
                (hashed, role, dept_id, username)
            )
    db.commit()
    
    db.execute("UPDATE departments SET head_id = 'u-head-central' WHERE id = 'dept-1'")
    db.execute("UPDATE departments SET head_id = 'u-head-dept' WHERE id = 'dept-2'")
    
    db.execute("INSERT OR IGNORE INTO settings (key, value) VALUES ('system_name', 'ЭДО ЛДПР')")
    db.execute("INSERT OR IGNORE INTO settings (key, value) VALUES ('company_name', 'ЛДПР')")
    db.commit()

# ============================================================
# МОДЕЛИ
# ============================================================

class UserModel:
    @staticmethod
    def get_by_id(uid):
        return get_db().execute("SELECT * FROM users WHERE uid = ?", (uid,)).fetchone()
    
    @staticmethod
    def get_by_username(username):
        return get_db().execute("SELECT * FROM users WHERE username = ?", (username,)).fetchone()
    
    @staticmethod
    def get_all():
        return get_db().execute("SELECT * FROM users WHERE is_active = 1 ORDER BY created_at DESC").fetchall()
    
    @staticmethod
    def create(uid, full_name, email, username, password, role, department_id):
        try:
            get_db().execute(
                "INSERT INTO users (uid, full_name, email, username, password, role, department_id) VALUES (?, ?, ?, ?, ?, ?, ?)",
                (uid, full_name, email, username, password, role, department_id)
            )
            get_db().commit()
            return True, None
        except sqlite3.IntegrityError as e:
            if 'username' in str(e):
                return False, 'Пользователь с таким логином уже существует'
            elif 'email' in str(e):
                return False, 'Пользователь с таким email уже существует'
            return False, 'Ошибка при создании пользователя'
    
    @staticmethod
    def update(uid, **kwargs):
        db = get_db()
        allowed = ['full_name', 'email', 'username', 'role', 'department_id', 'is_active']
        updates = {k: v for k, v in kwargs.items() if k in allowed and v is not None}
        if not updates:
            return
        set_clause = ', '.join(f"{k} = ?" for k in updates)
        values = list(updates.values()) + [uid]
        db.execute(f"UPDATE users SET {set_clause} WHERE uid = ?", values)
        db.commit()
    
    @staticmethod
    def delete(uid):
        get_db().execute("DELETE FROM users WHERE uid = ?", (uid,))
        get_db().commit()
    
    @staticmethod
    def get_by_department(department_id):
        return get_db().execute("SELECT * FROM users WHERE department_id = ? AND is_active = 1", (department_id,)).fetchall()
    
    @staticmethod
    def get_by_role(role):
        return get_db().execute("SELECT * FROM users WHERE role = ? AND is_active = 1", (role,)).fetchall()
    
    @staticmethod
    def get_stats():
        db = get_db()
        total = db.execute("SELECT COUNT(*) FROM users WHERE is_active = 1").fetchone()[0]
        by_role = {}
        roles = ['admin', 'secretary', 'head_central', 'head_department', 'assistant', 'executor']
        for role in roles:
            count = db.execute("SELECT COUNT(*) FROM users WHERE role = ? AND is_active = 1", (role,)).fetchone()[0]
            by_role[role] = count
        return {'total': total, 'by_role': by_role}

class DepartmentModel:
    @staticmethod
    def get_all():
        return get_db().execute("SELECT d.*, u.full_name as head_name FROM departments d LEFT JOIN users u ON d.head_id = u.uid ORDER BY d.name").fetchall()
    
    @staticmethod
    def get_by_id(dept_id):
        return get_db().execute("SELECT d.*, u.full_name as head_name FROM departments d LEFT JOIN users u ON d.head_id = u.uid WHERE d.id = ?", (dept_id,)).fetchone()
    
    @staticmethod
    def create(dept_id, name, description=None):
        try:
            get_db().execute("INSERT INTO departments (id, name, description) VALUES (?, ?, ?)", (dept_id, name, description))
            get_db().commit()
            return True
        except:
            return False
    
    @staticmethod
    def update(dept_id, **kwargs):
        db = get_db()
        allowed = ['name', 'description', 'head_id']
        updates = {k: v for k, v in kwargs.items() if k in allowed and v is not None}
        if not updates:
            return
        set_clause = ', '.join(f"{k} = ?" for k in updates)
        values = list(updates.values()) + [dept_id]
        db.execute(f"UPDATE departments SET {set_clause} WHERE id = ?", values)
        db.commit()
    
    @staticmethod
    def delete(dept_id):
        db = get_db()
        db.execute("UPDATE users SET department_id = NULL WHERE department_id = ?", (dept_id,))
        db.execute("DELETE FROM departments WHERE id = ?", (dept_id,))
        db.commit()
    
    @staticmethod
    def get_stats():
        total = get_db().execute("SELECT COUNT(*) FROM departments").fetchone()[0]
        return {'total': total}

class OrderModel:
    STATUSES = ['Черновик', 'На утверждении', 'Утверждено', 'В отделе', 'Назначен исполнитель',
                 'В работе', 'Готово к проверке', 'Подтверждено', 'На доработке', 'Закрыто', 'Отклонено']
    PRIORITIES = ['Низкий', 'Нормальный', 'Высокий', 'Срочный']
    
    @staticmethod
    def get_all():
        return get_db().execute("SELECT * FROM orders ORDER BY created_at DESC").fetchall()
    
    @staticmethod
    def get_by_id(order_id):
        order = get_db().execute("SELECT * FROM orders WHERE id = ?", (order_id,)).fetchone()
        if order:
            order = dict(order)
            if order.get('result'):
                try:
                    order['result'] = json.loads(order['result'])
                except:
                    pass
        return order
    
    @staticmethod
    def get_by_department(dept_id):
        return get_db().execute("SELECT * FROM orders WHERE assigned_department_id = ? ORDER BY created_at DESC", (dept_id,)).fetchall()
    
    @staticmethod
    def create(order_id, title, content, priority, status, created_by, creator_name, deadline=None, order_file=None):
        db = get_db()
        db.execute(
            "INSERT INTO orders (id, title, content, priority, status, created_by, creator_name, deadline, order_file, updated_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)",
            (order_id, title, content, priority, status, created_by, creator_name, deadline, order_file)
        )
        db.commit()
    
    @staticmethod
    def update(order_id, **kwargs):
        db = get_db()
        if 'result' in kwargs and kwargs['result']:
            kwargs['result'] = json.dumps(kwargs['result'], ensure_ascii=False)
        allowed = ['title', 'content', 'priority', 'status', 'assigned_department_id', 'assigned_executor_id', 'deadline', 'result', 'result_file', 'order_file', 'rejection_reason']
        updates = {k: v for k, v in kwargs.items() if k in allowed and v is not None}
        if not updates:
            return
        set_clause = ', '.join(f"{k} = ?" for k in updates) + ', updated_at = CURRENT_TIMESTAMP'
        values = list(updates.values()) + [order_id]
        db.execute(f"UPDATE orders SET {set_clause} WHERE id = ?", values)
        db.commit()
    
    @staticmethod
    def get_by_user(uid, role, department_id=None):
        db = get_db()
        if role == 'admin':
            return db.execute("SELECT * FROM orders ORDER BY created_at DESC").fetchall()
        elif role == 'assistant':
            return db.execute("SELECT * FROM orders WHERE created_by = ? ORDER BY created_at DESC", (uid,)).fetchall()
        elif role == 'head_central':
            return db.execute("SELECT * FROM orders WHERE status != 'Черновик' ORDER BY created_at DESC").fetchall()
        elif role == 'secretary':
            return db.execute("SELECT * FROM orders WHERE status = 'Утверждено' ORDER BY created_at DESC").fetchall()
        elif role == 'head_department' and department_id:
            return db.execute("SELECT * FROM orders WHERE assigned_department_id = ? ORDER BY created_at DESC", (department_id,)).fetchall()
        elif role == 'executor':
            return db.execute("SELECT * FROM orders WHERE assigned_executor_id = ? ORDER BY created_at DESC", (uid,)).fetchall()
        else:
            return db.execute("SELECT * FROM orders ORDER BY created_at DESC").fetchall()
    
    @staticmethod
    def get_stats(uid=None, role=None, department_id=None):
        orders = OrderModel.get_by_user(uid, role, department_id) if uid else OrderModel.get_all()
        stats = {'total': len(orders), 'pending': 0, 'approved': 0, 'in_work': 0, 'completed': 0, 'rejected': 0, 'draft': 0}
        for o in orders:
            if o['status'] == 'На утверждении':
                stats['pending'] += 1
            elif o['status'] == 'Утверждено':
                stats['approved'] += 1
            elif o['status'] == 'В работе':
                stats['in_work'] += 1
            elif o['status'] in ['Подтверждено', 'Закрыто']:
                stats['completed'] += 1
            elif o['status'] == 'Отклонено':
                stats['rejected'] += 1
            elif o['status'] == 'Черновик':
                stats['draft'] += 1
        return stats
    
    @staticmethod
    def get_global_stats():
        db = get_db()
        total = db.execute("SELECT COUNT(*) FROM orders").fetchone()[0]
        by_status = {}
        for status in OrderModel.STATUSES:
            count = db.execute("SELECT COUNT(*) FROM orders WHERE status = ?", (status,)).fetchone()[0]
            by_status[status] = count
        return {'total': total, 'by_status': by_status}

class OrderHistoryModel:
    @staticmethod
    def get_by_order(order_id):
        return get_db().execute("SELECT * FROM order_history WHERE order_id = ? ORDER BY created_at DESC", (order_id,)).fetchall()
    
    @staticmethod
    def add(order_id, action, user_name, user_role, details=None):
        get_db().execute("INSERT INTO order_history (order_id, action, user_name, user_role, details) VALUES (?, ?, ?, ?, ?)",
                        (order_id, action, user_name, user_role, details))
        get_db().commit()

class NotificationModel:
    @staticmethod
    def create(user_id, message, order_id=None):
        get_db().execute("INSERT INTO notifications (user_id, message, order_id) VALUES (?, ?, ?)", (user_id, message, order_id))
        get_db().commit()
    
    @staticmethod
    def get_by_user(user_id, limit=50):
        return get_db().execute("SELECT * FROM notifications WHERE user_id = ? ORDER BY created_at DESC LIMIT ?", (user_id, limit)).fetchall()
    
    @staticmethod
    def get_unread_count(user_id):
        return get_db().execute("SELECT COUNT(*) FROM notifications WHERE user_id = ? AND is_read = 0", (user_id,)).fetchone()[0]
    
    @staticmethod
    def mark_as_read(notification_id):
        get_db().execute("UPDATE notifications SET is_read = 1 WHERE id = ?", (notification_id,))
        get_db().commit()
    
    @staticmethod
    def mark_all_as_read(user_id):
        get_db().execute("UPDATE notifications SET is_read = 1 WHERE user_id = ?", (user_id,))
        get_db().commit()

class FileModel:
    @staticmethod
    def save(order_id, filename, filepath, filetype, filesize, uploaded_by):
        get_db().execute("INSERT INTO files (order_id, filename, filepath, filetype, filesize, uploaded_by) VALUES (?, ?, ?, ?, ?, ?)",
                        (order_id, filename, filepath, filetype, filesize, uploaded_by))
        get_db().commit()
        return get_db().execute("SELECT last_insert_rowid()").fetchone()[0]
    
    @staticmethod
    def get_by_order(order_id):
        return get_db().execute("SELECT * FROM files WHERE order_id = ?", (order_id,)).fetchall()
    
    @staticmethod
    def get_by_id(file_id):
        return get_db().execute("SELECT * FROM files WHERE id = ?", (file_id,)).fetchone()
    
    @staticmethod
    def delete(file_id):
        file = FileModel.get_by_id(file_id)
        if file and os.path.exists(file['filepath']):
            os.remove(file['filepath'])
        get_db().execute("DELETE FROM files WHERE id = ?", (file_id,))
        get_db().commit()

class SettingsModel:
    @staticmethod
    def get(key, default=None):
        result = get_db().execute("SELECT value FROM settings WHERE key = ?", (key,)).fetchone()
        return result['value'] if result else default
    
    @staticmethod
    def set(key, value):
        get_db().execute("INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)", (key, value))
        get_db().commit()

# ============================================================
# ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ
# ============================================================

def save_uploaded_file(file, order_id, uploaded_by):
    if not file or file.filename == '':
        return None
    filename = secure_filename(file.filename)
    unique_filename = f"{order_id}_{int(time.time())}_{filename}"
    filepath = os.path.join(UPLOAD_FOLDER, unique_filename)
    file.save(filepath)
    file_size = os.path.getsize(filepath)
    file_type = filename.rsplit('.', 1)[1].lower() if '.' in filename else 'unknown'
    FileModel.save(order_id, filename, filepath, file_type, file_size, uploaded_by)
    return unique_filename

def export_orders_to_excel(orders):
    if not EXCEL_AVAILABLE:
        return None
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Распоряжения"
    headers = ['№', 'ID', 'Название', 'Статус', 'Приоритет', 'Автор', 'Отдел', 'Исполнитель', 'Срок', 'Создано', 'Результат']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="003399", end_color="003399", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    for row, order in enumerate(orders, 2):
        ws.cell(row=row, column=1, value=row-1)
        ws.cell(row=row, column=2, value=order['id'])
        ws.cell(row=row, column=3, value=order['title'])
        ws.cell(row=row, column=4, value=order['status'])
        ws.cell(row=row, column=5, value=order['priority'])
        ws.cell(row=row, column=6, value=order['creator_name'])
        ws.cell(row=row, column=7, value=order['assigned_department_id'] or '-')
        ws.cell(row=row, column=8, value=order['assigned_executor_id'] or '-')
        ws.cell(row=row, column=9, value=order['deadline'] or '-')
        ws.cell(row=row, column=10, value=order['created_at'][:10] if order['created_at'] else '-')
        ws.cell(row=row, column=11, value='Выполнено' if order['result'] else '-')
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def export_users_to_excel(users):
    if not EXCEL_AVAILABLE:
        return None
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Пользователи"
    headers = ['№', 'UID', 'ФИО', 'Логин', 'Email', 'Роль', 'Отдел', 'Активен', 'Создан']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="003399", end_color="003399", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    for row, user in enumerate(users, 2):
        ws.cell(row=row, column=1, value=row-1)
        ws.cell(row=row, column=2, value=user['uid'])
        ws.cell(row=row, column=3, value=user['full_name'])
        ws.cell(row=row, column=4, value=user['username'])
        ws.cell(row=row, column=5, value=user['email'])
        ws.cell(row=row, column=6, value=user['role'])
        ws.cell(row=row, column=7, value=user['department_id'] or '-')
        ws.cell(row=row, column=8, value='Да' if user.get('is_active', 1) else 'Нет')
        ws.cell(row=row, column=9, value=user['created_at'][:10] if user['created_at'] else '-')
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# ============================================================
# ШАБЛОНЫ
# ============================================================

LOGIN_TEMPLATE = '''<!DOCTYPE html>
<html lang="ru">
<head>
    <meta charset="UTF-8">
    <title>Вход - ЭДО ЛДПР</title>
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/css/bootstrap.min.css" rel="stylesheet">
    <style>
        body { background: linear-gradient(135deg, #003399, #001a4d); min-height: 100vh; display: flex; align-items: center; justify-content: center; }
        .login-card { background: white; border-radius: 20px; padding: 40px; max-width: 400px; width: 100%; box-shadow: 0 10px 40px rgba(0,0,0,0.3); }
        .logo { text-align: center; font-size: 28px; font-weight: 900; color: #003399; }
        .btn-login { background: #003399; color: white; width: 100%; padding: 12px; border: none; border-radius: 10px; font-weight: bold; }
        .btn-login:hover { background: #002266; }
    </style>
</head>
<body>
    <div class="login-card">
        <div class="logo mb-3">{{ company_name }}</div>
        <div class="text-center mb-4 small text-muted">{{ system_name }}</div>
        {% with messages = get_flashed_messages(with_categories=true) %}
            {% if messages %}
                {% for cat, msg in messages %}
                    <div class="alert alert-{{ cat }}">{{ msg }}</div>
                {% endfor %}
            {% endif %}
        {% endwith %}
        <form method="POST">
            <input type="text" name="username" class="form-control mb-3" placeholder="Логин" required autofocus>
            <input type="password" name="password" class="form-control mb-3" placeholder="Пароль" required>
            <button type="submit" class="btn-login">Войти</button>
        </form>
        <div class="text-center mt-3 small text-muted">
            <strong>Тестовые учетные записи:</strong><br>
            admin / admin123 - Администратор<br>
            assistant / ast123 - Помощник<br>
            executor / exec123 - Исполнитель
        </div>
    </div>
</body>
</html>'''

DASHBOARD_TEMPLATE = '''<!DOCTYPE html>
<html lang="ru">
<head>
    <meta charset="UTF-8">
    <title>Рабочий стол - ЭДО ЛДПР</title>
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/css/bootstrap.min.css" rel="stylesheet">
    <link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.11.3/font/bootstrap-icons.css">
    <style>
        .navbar { background: #003399; }
        .navbar-brand, .navbar-text { color: white; }
        .sidebar { background: white; min-height: 100vh; }
        .sidebar .nav-link { color: #333; padding: 10px 15px; margin: 5px 0; border-radius: 10px; }
        .sidebar .nav-link:hover { background: #eef; color: #003399; }
        .sidebar .nav-link.active { background: #003399; color: white !important; }
        .card { border-radius: 15px; box-shadow: 0 2px 10px rgba(0,0,0,0.05); }
        .stat-card { border-left: 4px solid #003399; cursor: pointer; transition: transform 0.2s; }
        .stat-card:hover { transform: translateY(-3px); }
        .btn-primary { background: #003399; border: none; }
        .btn-primary:hover { background: #002266; }
    </style>
</head>
<body>
<nav class="navbar navbar-dark">
    <div class="container-fluid px-4">
        <a class="navbar-brand" href="/"><i class="bi bi-building"></i> ЭДО ЛДПР</a>
        <div class="d-flex align-items-center gap-3">
            <span class="navbar-text"><i class="bi bi-person-circle"></i> {{ current_user.full_name }}</span>
            <a href="/logout" class="btn btn-outline-light btn-sm"><i class="bi bi-box-arrow-right"></i> Выход</a>
        </div>
    </div>
</nav>
<div class="container-fluid">
    <div class="row">
        <div class="col-md-2 sidebar p-3">
            <div class="text-center mb-4">
                <div class="bg-primary text-white rounded-circle d-flex align-items-center justify-content-center mx-auto mb-2" style="width:60px;height:60px;font-size:24px;">{{ current_user.full_name[0] }}</div>
                <h6 class="mb-0">{{ current_user.full_name }}</h6>
                <small class="text-muted">{{ current_user.role }}</small>
            </div>
            <nav class="nav flex-column">
                <a class="nav-link active" href="/"><i class="bi bi-speedometer2 me-2"></i>Рабочий стол</a>
                <a class="nav-link" href="/orders"><i class="bi bi-file-text me-2"></i>Распоряжения</a>
                {% if current_user.role in ['head_department', 'admin'] %}
                <a class="nav-link" href="/department"><i class="bi bi-people me-2"></i>Отдел</a>
                {% endif %}
                {% if current_user.role == 'admin' %}
                <a class="nav-link" href="/admin"><i class="bi bi-gear me-2"></i>Администрирование</a>
                <a class="nav-link" href="/admin/stats"><i class="bi bi-graph-up me-2"></i>Статистика</a>
                {% endif %}
            </nav>
        </div>
        <div class="col-md-10 p-4">
            <h2 class="mb-4"><i class="bi bi-speedometer2 me-2"></i>Рабочий стол</h2>
            
            <div class="row mb-4">
                <div class="col-md-3 mb-3">
                    <div class="card stat-card p-3" onclick="location.href='/orders'">
                        <div class="d-flex justify-content-between align-items-center">
                            <div>
                                <small class="text-muted text-uppercase">Всего</small>
                                <h2 class="mb-0 mt-1">{{ stats.total }}</h2>
                            </div>
                            <i class="bi bi-files fs-1 text-primary opacity-50"></i>
                        </div>
                    </div>
                </div>
                <div class="col-md-3 mb-3">
                    <div class="card stat-card p-3" style="border-left-color:#f59e0b" onclick="location.href='/orders?status=На утверждении'">
                        <div class="d-flex justify-content-between align-items-center">
                            <div>
                                <small class="text-muted text-uppercase">На утверждении</small>
                                <h2 class="mb-0 mt-1">{{ stats.pending }}</h2>
                            </div>
                            <i class="bi bi-clock-history fs-1 text-warning opacity-50"></i>
                        </div>
                    </div>
                </div>
                <div class="col-md-3 mb-3">
                    <div class="card stat-card p-3" style="border-left-color:#10b981" onclick="location.href='/orders?status=Утверждено'">
                        <div class="d-flex justify-content-between align-items-center">
                            <div>
                                <small class="text-muted text-uppercase">Утверждено</small>
                                <h2 class="mb-0 mt-1">{{ stats.approved }}</h2>
                            </div>
                            <i class="bi bi-check-circle fs-1 text-success opacity-50"></i>
                        </div>
                    </div>
                </div>
                <div class="col-md-3 mb-3">
                    <div class="card stat-card p-3" style="border-left-color:#6366f1" onclick="location.href='/orders?status=В работе'">
                        <div class="d-flex justify-content-between align-items-center">
                            <div>
                                <small class="text-muted text-uppercase">В работе</small>
                                <h2 class="mb-0 mt-1">{{ stats.in_work }}</h2>
                            </div>
                            <i class="bi bi-play-circle fs-1 text-info opacity-50"></i>
                        </div>
                    </div>
                </div>
            </div>
            
            <div class="card">
                <div class="card-header bg-white d-flex justify-content-between align-items-center py-3">
                    <h5 class="mb-0 fw-bold"><i class="bi bi-clock-history me-2"></i>Последние распоряжения</h5>
                    <a href="/orders" class="btn btn-primary btn-sm">Все распоряжения <i class="bi bi-arrow-right"></i></a>
                </div>
                <div class="card-body p-0">
                    {% if orders %}
                    <div class="table-responsive">
                        <table class="table table-hover mb-0">
                            <thead class="table-light">
                                <tr>
                                    <th>Номер</th>
                                    <th>Название</th>
                                    <th>Статус</th>
                                    <th>Приоритет</th>
                                    <th>Создано</th>
                                </tr>
                            </thead>
                            <tbody>
                                {% for o in orders %}
                                <tr onclick="location.href='/orders/{{ o.id }}'" style="cursor:pointer">
                                    <td><small class="text-muted">#{{ o.id[:8] }}</small></d>
                                    <td><strong>{{ o.title }}</strong><br><small class="text-muted">{{ o.creator_name }}</small></d>
                                    <td><span class="badge bg-primary">{{ o.status }}</span></d>
                                    <td>{{ o.priority }}</d>
                                    <td><small>{{ o.created_at[:16] if o.created_at else '' }}</small></d>
                                </tr>
                                {% endfor %}
                            </tbody>
                        </table>
                    </div>
                    {% else %}
                    <div class="text-center py-5">
                        <i class="bi bi-inbox fs-1 text-muted"></i>
                        <p class="text-muted mt-2">Нет распоряжений</p>
                    </div>
                    {% endif %}
                </div>
            </div>
        </div>
    </div>
</div>
</body>
</html>'''

ORDERS_TEMPLATE = '''<!DOCTYPE html>
<html lang="ru">
<head>
    <meta charset="UTF-8">
    <title>Распоряжения - ЭДО ЛДПР</title>
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/css/bootstrap.min.css" rel="stylesheet">
    <link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.11.3/font/bootstrap-icons.css">
    <style>
        .navbar { background: #003399; }
        .navbar-brand, .navbar-text { color: white; }
        .sidebar { background: white; min-height: 100vh; }
        .sidebar .nav-link { color: #333; padding: 10px 15px; margin: 5px 0; border-radius: 10px; }
        .sidebar .nav-link:hover { background: #eef; color: #003399; }
        .sidebar .nav-link.active { background: #003399; color: white !important; }
        .btn-primary { background: #003399; border: none; }
        .btn-primary:hover { background: #002266; }
        .btn-create { background: #003399; color: white; padding: 10px 24px; border: none; border-radius: 8px; font-weight: bold; }
        .btn-create:hover { background: #002266; }
    </style>
</head>
<body>
<nav class="navbar navbar-dark">
    <div class="container-fluid px-4">
        <a class="navbar-brand" href="/"><i class="bi bi-building"></i> ЭДО ЛДПР</a>
        <div class="d-flex align-items-center gap-3">
            <span class="navbar-text"><i class="bi bi-person-circle"></i> {{ current_user.full_name }}</span>
            <a href="/logout" class="btn btn-outline-light btn-sm"><i class="bi bi-box-arrow-right"></i> Выход</a>
        </div>
    </div>
</nav>
<div class="container-fluid">
    <div class="row">
        <div class="col-md-2 sidebar p-3">
            <div class="text-center mb-4">
                <div class="bg-primary text-white rounded-circle d-flex align-items-center justify-content-center mx-auto mb-2" style="width:60px;height:60px;font-size:24px;">{{ current_user.full_name[0] }}</div>
                <h6 class="mb-0">{{ current_user.full_name }}</h6>
                <small class="text-muted">{{ current_user.role }}</small>
            </div>
            <nav class="nav flex-column">
                <a class="nav-link" href="/"><i class="bi bi-speedometer2 me-2"></i>Рабочий стол</a>
                <a class="nav-link active" href="/orders"><i class="bi bi-file-text me-2"></i>Распоряжения</a>
                {% if current_user.role in ['head_department', 'admin'] %}
                <a class="nav-link" href="/department"><i class="bi bi-people me-2"></i>Отдел</a>
                {% endif %}
                {% if current_user.role == 'admin' %}
                <a class="nav-link" href="/admin"><i class="bi bi-gear me-2"></i>Администрирование</a>
                <a class="nav-link" href="/admin/stats"><i class="bi bi-graph-up me-2"></i>Статистика</a>
                {% endif %}
            </nav>
        </div>
        <div class="col-md-10 p-4">
            <div class="d-flex justify-content-between align-items-center mb-4">
                <h2 class="mb-0"><i class="bi bi-file-text me-2"></i>Распоряжения</h2>
                {% if current_user.role == 'assistant' %}
                <a href="/order/create" class="btn btn-create"><i class="bi bi-plus-lg me-2"></i>Создать распоряжение</a>
                {% endif %}
            </div>
            
            <div class="card">
                <div class="card-body p-0">
                    <div class="table-responsive">
                        <table class="table table-hover mb-0">
                            <thead class="table-light">
                                <tr>
                                    <th>№</th>
                                    <th>Название</th>
                                    <th>Статус</th>
                                    <th>Приоритет</th>
                                    <th>Автор</th>
                                    <th>Создано</th>
                                </tr>
                            </thead>
                            <tbody>
                                {% for o in orders %}
                                <tr onclick="location.href='/orders/{{ o.id }}'" style="cursor:pointer">
                                    <td><small class="text-muted">#{{ o.id[:8] }}</small></d>
                                    <td><strong>{{ o.title }}</strong></d>
                                    <td><span class="badge bg-primary">{{ o.status }}</span></d>
                                    <td>{{ o.priority }}</d>
                                    <td>{{ o.creator_name }}</d>
                                    <td><small>{{ o.created_at[:16] if o.created_at else '' }}</small></d>
                                </tr>
                                {% endfor %}
                            </tbody>
                        </table>
                    </div>
                </div>
            </div>
        </div>
    </div>
</div>
</body>
</html>'''

CREATE_ORDER_TEMPLATE = '''<!DOCTYPE html>
<html lang="ru">
<head>
    <meta charset="UTF-8">
    <title>Создание распоряжения - ЭДО ЛДПР</title>
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/css/bootstrap.min.css" rel="stylesheet">
    <link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.11.3/font/bootstrap-icons.css">
    <style>
        .navbar { background: #003399; }
        .navbar-brand, .navbar-text { color: white; }
        .sidebar { background: white; min-height: 100vh; }
        .sidebar .nav-link { color: #333; padding: 10px 15px; margin: 5px 0; border-radius: 10px; }
        .sidebar .nav-link:hover { background: #eef; color: #003399; }
        .btn-primary { background: #003399; border: none; }
        .btn-primary:hover { background: #002266; }
        .btn-outline-primary { border-color: #003399; color: #003399; }
        .btn-outline-primary:hover { background: #003399; color: white; }
        .form-control:focus, .form-select:focus { border-color: #003399; box-shadow: 0 0 0 0.2rem rgba(0,51,153,0.25); }
    </style>
</head>
<body>
<nav class="navbar navbar-dark">
    <div class="container-fluid px-4">
        <a class="navbar-brand" href="/"><i class="bi bi-building"></i> ЭДО ЛДПР</a>
        <div class="d-flex align-items-center gap-3">
            <span class="navbar-text"><i class="bi bi-person-circle"></i> {{ current_user.full_name }}</span>
            <a href="/logout" class="btn btn-outline-light btn-sm"><i class="bi bi-box-arrow-right"></i> Выход</a>
        </div>
    </div>
</nav>
<div class="container-fluid">
    <div class="row">
        <div class="col-md-2 sidebar p-3">
            <div class="text-center mb-4">
                <div class="bg-primary text-white rounded-circle d-flex align-items-center justify-content-center mx-auto mb-2" style="width:60px;height:60px;font-size:24px;">{{ current_user.full_name[0] }}</div>
                <h6 class="mb-0">{{ current_user.full_name }}</h6>
                <small class="text-muted">{{ current_user.role }}</small>
            </div>
            <nav class="nav flex-column">
                <a class="nav-link" href="/"><i class="bi bi-speedometer2 me-2"></i>Рабочий стол</a>
                <a class="nav-link active" href="/orders"><i class="bi bi-file-text me-2"></i>Распоряжения</a>
                {% if current_user.role in ['head_department', 'admin'] %}
                <a class="nav-link" href="/department"><i class="bi bi-people me-2"></i>Отдел</a>
                {% endif %}
                {% if current_user.role == 'admin' %}
                <a class="nav-link" href="/admin"><i class="bi bi-gear me-2"></i>Администрирование</a>
                {% endif %}
            </nav>
        </div>
        <div class="col-md-10 p-4">
            <div class="d-flex align-items-center mb-4">
                <a href="/orders" class="btn btn-outline-secondary btn-sm me-3"><i class="bi bi-arrow-left"></i> Назад</a>
                <h2 class="mb-0"><i class="bi bi-file-earmark-plus me-2"></i>Создание распоряжения</h2>
            </div>
            
            <div class="card">
                <div class="card-body p-4">
                    <form method="POST" action="/order/create" enctype="multipart/form-data">
                        <div class="mb-3">
                            <label class="form-label fw-bold">Заголовок <span class="text-danger">*</span></label>
                            <input type="text" name="title" class="form-control form-control-lg" placeholder="Введите заголовок распоряжения" required>
                        </div>
                        
                        <div class="row mb-3">
                            <div class="col-md-6">
                                <label class="form-label fw-bold">Приоритет</label>
                                <select name="priority" class="form-select">
                                    <option value="Низкий">Низкий</option>
                                    <option value="Нормальный" selected>Нормальный</option>
                                    <option value="Высокий">Высокий</option>
                                    <option value="Срочный">Срочный</option>
                                </select>
                            </div>
                            <div class="col-md-6">
                                <label class="form-label fw-bold">Срок исполнения</label>
                                <input type="date" name="deadline" class="form-control">
                            </div>
                        </div>
                        
                        <div class="mb-3">
                            <label class="form-label fw-bold">Содержание <span class="text-danger">*</span></label>
                            <textarea name="content" class="form-control" rows="8" placeholder="Введите текст распоряжения..." required></textarea>
                        </div>
                        
                        <div class="mb-4">
                            <label class="form-label fw-bold"><i class="bi bi-paperclip me-1"></i> Прикрепить файл</label>
                            <input type="file" name="order_file" class="form-control">
                            <small class="text-muted">Поддерживаемые форматы: PDF, DOC, DOCX, XLS, XLSX, JPG, PNG, ZIP (максимум 16MB)</small>
                        </div>
                        
                        <hr>
                        
                        <div class="d-flex justify-content-end gap-2">
                            <a href="/orders" class="btn btn-secondary">Отмена</a>
                            <button type="submit" name="action" value="draft" class="btn btn-outline-primary">
                                <i class="bi bi-save"></i> Сохранить черновик
                            </button>
                            <button type="submit" name="action" value="submit" class="btn btn-primary">
                                <i class="bi bi-send"></i> Отправить на утверждение
                            </button>
                        </div>
                    </form>
                </div>
            </div>
        </div>
    </div>
</div>
</body>
</html>'''

ORDER_DETAILS_TEMPLATE = '''<!DOCTYPE html>
<html lang="ru">
<head>
    <meta charset="UTF-8">
    <title>{{ order.title }} - ЭДО ЛДПР</title>
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/css/bootstrap.min.css" rel="stylesheet">
    <link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.11.3/font/bootstrap-icons.css">
    <style>
        .navbar { background: #003399; }
        .navbar-brand, .navbar-text { color: white; }
        .sidebar { background: white; min-height: 100vh; }
        .sidebar .nav-link { color: #333; padding: 10px 15px; margin: 5px 0; border-radius: 10px; }
        .sidebar .nav-link:hover { background: #eef; color: #003399; }
        .btn-primary { background: #003399; border: none; }
        .btn-primary:hover { background: #002266; }
        .status-badge { font-size: 1rem; padding: 8px 20px; border-radius: 30px; }
    </style>
</head>
<body>
<nav class="navbar navbar-dark">
    <div class="container-fluid px-4">
        <a class="navbar-brand" href="/"><i class="bi bi-building"></i> ЭДО ЛДПР</a>
        <div class="d-flex align-items-center gap-3">
            <span class="navbar-text"><i class="bi bi-person-circle"></i> {{ current_user.full_name }}</span>
            <a href="/logout" class="btn btn-outline-light btn-sm"><i class="bi bi-box-arrow-right"></i> Выход</a>
        </div>
    </div>
</nav>
<div class="container-fluid">
    <div class="row">
        <div class="col-md-2 sidebar p-3">
            <div class="text-center mb-4">
                <div class="bg-primary text-white rounded-circle d-flex align-items-center justify-content-center mx-auto mb-2" style="width:60px;height:60px;font-size:24px;">{{ current_user.full_name[0] }}</div>
                <h6 class="mb-0">{{ current_user.full_name }}</h6>
                <small class="text-muted">{{ current_user.role }}</small>
            </div>
            <nav class="nav flex-column">
                <a class="nav-link" href="/"><i class="bi bi-speedometer2 me-2"></i>Рабочий стол</a>
                <a class="nav-link active" href="/orders"><i class="bi bi-file-text me-2"></i>Распоряжения</a>
                {% if current_user.role in ['head_department', 'admin'] %}
                <a class="nav-link" href="/department"><i class="bi bi-people me-2"></i>Отдел</a>
                {% endif %}
                {% if current_user.role == 'admin' %}
                <a class="nav-link" href="/admin"><i class="bi bi-gear me-2"></i>Администрирование</a>
                {% endif %}
            </nav>
        </div>
        <div class="col-md-10 p-4">
            <div class="d-flex justify-content-between align-items-center mb-3">
                <a href="/orders" class="btn btn-outline-secondary btn-sm"><i class="bi bi-arrow-left"></i> Назад</a>
                <div class="dropdown">
                    <button class="btn btn-outline-secondary btn-sm dropdown-toggle" data-bs-toggle="dropdown">
                        <i class="bi bi-three-dots"></i> Действия
                    </button>
                    <ul class="dropdown-menu dropdown-menu-end">
                        <li><a class="dropdown-item" href="#" onclick="window.print()"><i class="bi bi-printer"></i> Печать</a></li>
                    </ul>
                </div>
            </div>
            
            <div class="card mb-4">
                <div class="card-body p-4">
                    <div class="d-flex justify-content-between align-items-start mb-3">
                        <div>
                            <span class="badge bg-secondary mb-2">№ {{ order.id }}</span>
                            <h2 class="fw-bold mb-0">{{ order.title }}</h2>
                        </div>
                        <span class="badge bg-primary status-badge">{{ order.status }}</span>
                    </div>
                    
                    <div class="row g-3 py-3 border-top border-bottom">
                        <div class="col-md-3">
                            <small class="text-muted d-block">Автор</small>
                            <strong>{{ order.creator_name }}</strong>
                        </div>
                        <div class="col-md-3">
                            <small class="text-muted d-block">Срок исполнения</small>
                            <strong>{{ order.deadline or 'Не указан' }}</strong>
                        </div>
                        <div class="col-md-3">
                            <small class="text-muted d-block">Приоритет</small>
                            <strong class="{% if order.priority == 'Срочный' %}text-danger{% endif %}">{{ order.priority }}</strong>
                        </div>
                        <div class="col-md-3">
                            <small class="text-muted d-block">Создано</small>
                            <strong>{{ order.created_at[:16] if order.created_at else '' }}</strong>
                        </div>
                    </div>
                    
                    <div class="mt-4">
                        <h6 class="fw-bold">Текст распоряжения</h6>
                        <div class="bg-light p-3 rounded" style="white-space: pre-wrap;">{{ order.content }}</div>
                    </div>
                    
                    {% if order.order_file %}
                    <div class="mt-3">
                        <h6 class="fw-bold"><i class="bi bi-paperclip"></i> Прикреплённый файл</h6>
                        <a href="/uploads/{{ order.order_file }}" class="btn btn-sm btn-outline-primary">
                            <i class="bi bi-download"></i> Скачать файл
                        </a>
                    </div>
                    {% endif %}
                    
                    {% if order.result %}
                    <div class="mt-4 p-3 bg-success bg-opacity-10 rounded">
                        <h6 class="text-success fw-bold"><i class="bi bi-check-circle-fill"></i> Результат выполнения</h6>
                        <div class="mt-2">{{ order.result.content }}</div>
                        {% if order.result_file %}
                        <a href="/uploads/{{ order.result_file }}" class="btn btn-sm btn-outline-success mt-2">
                            <i class="bi bi-download"></i> Скачать файл результата
                        </a>
                        {% endif %}
                        <small class="text-muted d-block mt-2">Подано: {{ order.result.submittedAt[:16] }}</small>
                    </div>
                    {% endif %}
                </div>
            </div>
            
            <div class="card">
                <div class="card-body p-4">
                    <h5 class="fw-bold mb-3"><i class="bi bi-gear"></i> Доступные действия</h5>
                    
                    <!-- 1. Руководитель ЦА - утверждение/отклонение -->
                    {% if current_user.role == 'head_central' and order.status == 'На утверждении' %}
                    <form method="POST" action="/orders/{{ order.id }}/status" class="mb-2">
                        <input type="hidden" name="status" value="Утверждено">
                        <button class="btn btn-success w-100 mb-2"><i class="bi bi-check-circle"></i> Утвердить распоряжение</button>
                    </form>
                    <form method="POST" action="/orders/{{ order.id }}/status">
                        <input type="hidden" name="status" value="Отклонено">
                        <textarea name="comment" class="form-control mb-2" rows="2" placeholder="Укажите причину отклонения..."></textarea>
                        <button class="btn btn-danger w-100"><i class="bi bi-x-circle"></i> Отклонить</button>
                    </form>
                    {% endif %}
                    
                    <!-- 2. Секретарь - назначение отдела -->
                    {% if current_user.role == 'secretary' and order.status == 'Утверждено' %}
                    <form method="POST" action="/orders/{{ order.id }}/status">
                        <input type="hidden" name="status" value="В отделе">
                        <label class="form-label fw-bold">Выберите отдел</label>
                        <select name="department_id" class="form-select mb-2" required>
                            <option value="">-- Выберите отдел --</option>
                            {% for d in departments %}
                            <option value="{{ d.id }}">{{ d.name }}</option>
                            {% endfor %}
                        </select>
                        <button class="btn btn-primary w-100"><i class="bi bi-building"></i> Назначить отдел</button>
                    </form>
                    {% endif %}
                    
                    <!-- 3. Начальник отдела - назначение исполнителя -->
                    {% if current_user.role == 'head_department' and order.status == 'В отделе' %}
                    <form method="POST" action="/orders/{{ order.id }}/status">
                        <input type="hidden" name="status" value="Назначен исполнитель">
                        <label class="form-label fw-bold">Выберите исполнителя</label>
                        <select name="executor_id" class="form-select mb-2" required>
                            <option value="">-- Выберите исполнителя --</option>
                            {% for u in dept_users %}
                            <option value="{{ u.uid }}">{{ u.full_name }}</option>
                            {% endfor %}
                        </select>
                        <button class="btn btn-primary w-100"><i class="bi bi-person-check"></i> Назначить исполнителя</button>
                    </form>
                    {% endif %}
                    
                    <!-- 4. Исполнитель - взять в работу -->
                    {% if current_user.role == 'executor' and order.status == 'Назначен исполнитель' and order.assigned_executor_id == current_user.uid %}
                    <form method="POST" action="/orders/{{ order.id }}/status">
                        <input type="hidden" name="status" value="В работе">
                        <button class="btn btn-primary w-100"><i class="bi bi-play-circle"></i> Взять в работу</button>
                    </form>
                    {% endif %}
                    
                    <!-- 5. Исполнитель - сдать работу -->
                    {% if current_user.role == 'executor' and order.status == 'В работе' and order.assigned_executor_id == current_user.uid %}
                    <form method="POST" action="/orders/{{ order.id }}/submit" enctype="multipart/form-data">
                        <label class="form-label fw-bold">Результат выполнения</label>
                        <textarea name="result_content" class="form-control mb-2" rows="4" placeholder="Опишите результат выполнения..." required></textarea>
                        <label class="form-label fw-bold"><i class="bi bi-paperclip"></i> Прикрепить файл</label>
                        <input type="file" name="result_file" class="form-control mb-2">
                        <button class="btn btn-success w-100"><i class="bi bi-check-circle"></i> Сдать работу на проверку</button>
                    </form>
                    {% endif %}
                    
                    <!-- 6. Начальник отдела - проверка результата -->
                    {% if current_user.role == 'head_department' and order.status == 'Готово к проверке' %}
                    <form method="POST" action="/orders/{{ order.id }}/status" class="mb-2">
                        <input type="hidden" name="status" value="Подтверждено">
                        <button class="btn btn-success w-100 mb-2"><i class="bi bi-check-circle"></i> Подтвердить выполнение</button>
                    </form>
                    <form method="POST" action="/orders/{{ order.id }}/status">
                        <input type="hidden" name="status" value="На доработке">
                        <textarea name="comment" class="form-control mb-2" rows="2" placeholder="Укажите причину доработки..."></textarea>
                        <button class="btn btn-warning w-100"><i class="bi bi-arrow-counterclockwise"></i> Отправить на доработку</button>
                    </form>
                    {% endif %}
                    
                    <!-- 7. Руководитель ЦА - закрытие -->
                    {% if current_user.role == 'head_central' and order.status == 'Подтверждено' %}
                    <form method="POST" action="/orders/{{ order.id }}/status">
                        <input type="hidden" name="status" value="Закрыто">
                        <button class="btn btn-success w-100"><i class="bi bi-check-circle"></i> Закрыть распоряжение</button>
                    </form>
                    {% endif %}
                </div>
            </div>
        </div>
    </div>
</div>
</body>
</html>'''

DEPARTMENT_TEMPLATE = '''<!DOCTYPE html>
<html lang="ru">
<head>
    <meta charset="UTF-8">
    <title>Отдел - ЭДО ЛДПР</title>
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/css/bootstrap.min.css" rel="stylesheet">
    <link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.11.3/font/bootstrap-icons.css">
    <style>.navbar { background: #003399; }.sidebar { background: white; min-height: 100vh; }</style>
</head>
<body>
<nav class="navbar navbar-dark"><div class="container-fluid"><a class="navbar-brand" href="/">ЭДО ЛДПР</a>
<div class="d-flex"><span class="navbar-text me-3">{{ current_user.full_name }}</span><a href="/logout" class="btn btn-outline-light btn-sm">Выход</a></div></div></nav>
<div class="container-fluid"><div class="row"><div class="col-md-2 sidebar p-3">
<div class="text-center mb-4"><div class="bg-primary text-white rounded-circle d-flex align-items-center justify-content-center mx-auto mb-2" style="width:60px;height:60px">{{ current_user.full_name[0] }}</div>
<h6>{{ current_user.full_name }}</h6><small class="text-muted">{{ current_user.role }}</small></div>
<nav class="nav flex-column"><a class="nav-link" href="/">Рабочий стол</a><a class="nav-link" href="/orders">Распоряжения</a>
{% if current_user.role in ['head_department','admin'] %}<a class="nav-link active" href="/department">Отдел</a>{% endif %}
{% if current_user.role == 'admin' %}<a class="nav-link" href="/admin">Админ</a>{% endif %}</nav>
</div><div class="col-md-10 p-4">
<h2><i class="bi bi-building me-2"></i>{{ department.name }}</h2>
<div class="row"><div class="col-md-6"><div class="card p-3"><h5>Сотрудники</h5><table class="table">{% for u in users %}<tr><td>{{ u.full_name }}</td><td>{{ u.role }}</td><td>{{ u.email }}</td></tr>{% endfor %}</table></div></div>
<div class="col-md-6"><div class="card p-3"><h5>Распоряжения отдела</h5>{% for o in orders %}<a href="/orders/{{ o.id }}" class="d-block mb-2">{{ o.title }} - {{ o.status }}</a>{% endfor %}</div></div></div>
</div></div></div>
</body></html>'''

ADMIN_TEMPLATE = '''<!DOCTYPE html>
<html lang="ru">
<head>
    <meta charset="UTF-8">
    <title>Администрирование - ЭДО ЛДПР</title>
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/css/bootstrap.min.css" rel="stylesheet">
    <link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.11.3/font/bootstrap-icons.css">
    <style>.navbar { background: #003399; }.sidebar { background: white; min-height: 100vh; }</style>
</head>
<body>
<nav class="navbar navbar-dark"><div class="container-fluid"><a class="navbar-brand" href="/">ЭДО ЛДПР</a>
<div class="d-flex"><span class="navbar-text me-3">{{ current_user.full_name }}</span><a href="/logout" class="btn btn-outline-light btn-sm">Выход</a></div></div></nav>
<div class="container-fluid"><div class="row"><div class="col-md-2 sidebar p-3">
<div class="text-center mb-4"><div class="bg-primary text-white rounded-circle d-flex align-items-center justify-content-center mx-auto mb-2" style="width:60px;height:60px">{{ current_user.full_name[0] }}</div>
<h6>{{ current_user.full_name }}</h6><small class="text-muted">{{ current_user.role }}</small></div>
<nav class="nav flex-column"><a class="nav-link" href="/">Рабочий стол</a><a class="nav-link" href="/orders">Распоряжения</a><a class="nav-link active" href="/admin">Админ</a><a class="nav-link" href="/admin/stats">Статистика</a></nav>
</div><div class="col-md-10 p-4">
<h2>Администрирование</h2>
<div class="card p-3"><div class="d-flex justify-content-between"><h5>Пользователи</h5><div><a href="/admin/export/users" class="btn btn-success btn-sm me-2">Excel</a><button class="btn btn-primary btn-sm" data-bs-toggle="modal" data-bs-target="#addUserModal">+ Добавить</button></div></div>
<table class="table table-hover"><thead class="table-light"><tr><th>ФИО</th><th>Логин</th><th>Роль</th><th>Email</th><th>Отдел</th><th>Действия</th></tr></thead>
<tbody>{% for u in users %}<tr><td>{{ u.full_name }}</td><td>{{ u.username }}</td><td>{{ u.role }}</td><td>{{ u.email }}</td><td>{{ u.department_id or '-' }}</td>
<td><a href="/admin/users/{{ u.uid }}/delete" class="btn btn-sm btn-outline-danger" onclick="return confirm('Удалить?')">Удалить</a></td></tr>{% endfor %}</tbody><td></div>
<div class="card p-3 mt-3"><h5>Отделы</h5><button class="btn btn-primary btn-sm mb-3" data-bs-toggle="modal" data-bs-target="#addDeptModal">+ Добавить отдел</button>
<table class="table"><thead class="table-light"><tr><th>Название</th><th>Руководитель</th><th>Действия</th></tr></thead>
<tbody>{% for d in departments %}<tr><td>{{ d.name }}</td><td>{{ d.head_name or 'Не назначен' }}</td><td><a href="/admin/departments/{{ d.id }}/delete" class="btn btn-sm btn-outline-danger" onclick="return confirm('Удалить отдел?')">Удалить</a></td></tr>{% endfor %}</tbody></table></div>
<div class="modal fade" id="addUserModal" tabindex="-1"><div class="modal-dialog"><div class="modal-content"><div class="modal-header"><h5>Добавить пользователя</h5><button type="button" class="btn-close" data-bs-dismiss="modal"></button></div>
<form method="POST" action="/admin/users/create"><div class="modal-body"><input type="text" name="full_name" class="form-control mb-2" placeholder="ФИО" required>
<input type="text" name="username" class="form-control mb-2" placeholder="Логин" required><input type="email" name="email" class="form-control mb-2" placeholder="Email" required>
<input type="password" name="password" class="form-control mb-2" placeholder="Пароль" required>
<select name="role" class="form-select mb-2"><option value="executor">Исполнитель</option><option value="assistant">Помощник</option><option value="head_department">Начальник отдела</option><option value="head_central">Руководитель ЦА</option><option value="secretary">Секретарь</option><option value="admin">Администратор</option></select>
<select name="department_id" class="form-select"><option value="">Без отдела</option>{% for d in departments %}<option value="{{ d.id }}">{{ d.name }}</option>{% endfor %}</select></div>
<div class="modal-footer"><button type="submit" class="btn btn-primary">Создать</button></div></form></div></div></div>
<div class="modal fade" id="addDeptModal" tabindex="-1"><div class="modal-dialog"><div class="modal-content"><div class="modal-header"><h5>Добавить отдел</h5><button type="button" class="btn-close" data-bs-dismiss="modal"></button></div>
<form method="POST" action="/admin/departments/create"><div class="modal-body"><input type="text" name="name" class="form-control mb-2" placeholder="Название" required>
<textarea name="description" class="form-control" placeholder="Описание"></textarea></div>
<div class="modal-footer"><button type="submit" class="btn btn-primary">Создать</button></div></form></div></div></div>
</div></div></div>
</body></html>'''

STATS_TEMPLATE = '''<!DOCTYPE html>
<html lang="ru">
<head>
    <meta charset="UTF-8">
    <title>Статистика - ЭДО ЛДПР</title>
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/css/bootstrap.min.css" rel="stylesheet">
    <link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.11.3/font/bootstrap-icons.css">
    <style>.navbar { background: #003399; }.sidebar { background: white; min-height: 100vh; }</style>
</head>
<body>
<nav class="navbar navbar-dark"><div class="container-fluid"><a class="navbar-brand" href="/">ЭДО ЛДПР</a>
<div class="d-flex"><span class="navbar-text me-3">{{ current_user.full_name }}</span><a href="/logout" class="btn btn-outline-light btn-sm">Выход</a></div></div></nav>
<div class="container-fluid"><div class="row"><div class="col-md-2 sidebar p-3">
<div class="text-center mb-4"><div class="bg-primary text-white rounded-circle d-flex align-items-center justify-content-center mx-auto mb-2" style="width:60px;height:60px">{{ current_user.full_name[0] }}</div>
<h6>{{ current_user.full_name }}</h6><small class="text-muted">{{ current_user.role }}</small></div>
<nav class="nav flex-column"><a class="nav-link" href="/">Рабочий стол</a><a class="nav-link" href="/orders">Распоряжения</a><a class="nav-link" href="/admin">Админ</a><a class="nav-link active" href="/admin/stats">Статистика</a></nav>
</div><div class="col-md-10 p-4">
<h2>Статистика системы</h2>
<div class="row mb-4"><div class="col-md-3"><div class="card p-3 text-center"><i class="bi bi-people fs-1 text-primary"></i><h3>{{ user_stats.total }}</h3><small>Пользователей</small></div></div>
<div class="col-md-3"><div class="card p-3 text-center"><i class="bi bi-building fs-1 text-success"></i><h3>{{ dept_stats.total }}</h3><small>Отделов</small></div></div>
<div class="col-md-3"><div class="card p-3 text-center"><i class="bi bi-file-text fs-1 text-warning"></i><h3>{{ order_stats.total }}</h3><small>Распоряжений</small></div></div>
<div class="col-md-3"><div class="card p-3 text-center"><i class="bi bi-check-circle fs-1 text-info"></i><h3>{{ completed_orders }}</h3><small>Выполнено</small></div></div></div>
<div class="card p-3"><h5>Распределение по ролям</h5><table class="table"><thead><tr><th>Роль</th><th>Количество</th></tr></thead>
<tbody>{% for role, count in user_stats.by_role.items() %}<tr><td>{% if role == 'admin' %}Администратор{% elif role == 'secretary' %}Секретарь{% elif role == 'head_central' %}Руководитель ЦА{% elif role == 'head_department' %}Начальник отдела{% elif role == 'assistant' %}Помощник{% else %}Исполнитель{% endif %}</td><td>{{ count }}</td></tr>{% endfor %}</tbody></table></div>
</div></div></div>
</body></html>'''

# ============================================================
# МАРШРУТЫ FLASK
# ============================================================

@app_flask.route('/login', methods=['GET', 'POST'])
def login():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        user = UserModel.get_by_username(username)
        if user and check_password_hash(user['password'], password):
            session.permanent = True
            session['user_id'] = user['uid']
            session['user_name'] = user['full_name']
            session['user_role'] = user['role']
            session['department_id'] = user['department_id']
            flash(f'Добро пожаловать, {user["full_name"]}!', 'success')
            return redirect(url_for('dashboard'))
        flash('Неверный логин или пароль', 'danger')
    return render_template_string(LOGIN_TEMPLATE,
        system_name=SettingsModel.get('system_name', 'ЭДО ЛДПР'),
        company_name=SettingsModel.get('company_name', 'ЛДПР'))

@app_flask.route('/logout')
def logout():
    session.clear()
    flash('Вы вышли из системы', 'info')
    return redirect(url_for('login'))

@app_flask.route('/')
@login_required
def dashboard():
    stats = OrderModel.get_stats(uid=session['user_id'], role=session['user_role'], department_id=session.get('department_id'))
    orders = OrderModel.get_by_user(uid=session['user_id'], role=session['user_role'], department_id=session.get('department_id'))[:5]
    return render_template_string(DASHBOARD_TEMPLATE, stats=stats, orders=orders)

@app_flask.route('/orders')
@login_required
def orders():
    all_orders = OrderModel.get_by_user(uid=session['user_id'], role=session['user_role'], department_id=session.get('department_id'))
    return render_template_string(ORDERS_TEMPLATE, orders=all_orders)

@app_flask.route('/order/create', methods=['GET', 'POST'])
@login_required
@role_required('assistant')
def create_order_page():
    if request.method == 'GET':
        return render_template_string(CREATE_ORDER_TEMPLATE)
    
    # POST - создание распоряжения
    title = request.form.get('title', '').strip()
    content = request.form.get('content', '').strip()
    action = request.form.get('action', 'submit')
    
    if not title or not content:
        flash('Заголовок и содержание обязательны', 'danger')
        return redirect(url_for('create_order_page'))
    
    order_id = 'ORD-' + str(uuid.uuid4())[:8].upper()
    is_draft = (action == 'draft')
    status = 'Черновик' if is_draft else 'На утверждении'
    user = UserModel.get_by_id(session['user_id'])
    
    order_file = None
    if 'order_file' in request.files:
        file = request.files['order_file']
        if file and file.filename and allowed_file(file.filename):
            order_file = save_uploaded_file(file, order_id, session['user_id'])
    
    OrderModel.create(order_id, title, content, request.form.get('priority', 'Нормальный'), 
                      status, session['user_id'], user['full_name'], 
                      request.form.get('deadline') or None, order_file)
    
    OrderHistoryModel.add(order_id, 'Создание распоряжения', user['full_name'], 
                          session['user_role'], f'Статус: {status}')
    
    if not is_draft:
        head_central_list = UserModel.get_by_role('head_central')
        for hc in head_central_list:
            NotificationModel.create(hc['uid'], f'Новое распоряжение на утверждение: {title}', order_id)
        flash('Распоряжение создано и отправлено на утверждение!', 'success')
    else:
        flash('Распоряжение сохранено как черновик', 'success')
    
    return redirect(url_for('orders'))

@app_flask.route('/orders/<order_id>')
@login_required
def order_details(order_id):
    order = OrderModel.get_by_id(order_id)
    if not order:
        flash('Распоряжение не найдено', 'danger')
        return redirect(url_for('orders'))
    
    # Проверка прав доступа
    user_role = session['user_role']
    user_id = session['user_id']
    user_dept = session.get('department_id')
    
    can_view = False
    if user_role == 'admin':
        can_view = True
    elif user_role == 'head_central':
        can_view = True
    elif user_role == 'assistant' and order['created_by'] == user_id:
        can_view = True
    elif user_role == 'secretary' and order['status'] == 'Утверждено':
        can_view = True
    elif user_role == 'head_department' and order['assigned_department_id'] == user_dept:
        can_view = True
    elif user_role == 'executor' and order['assigned_executor_id'] == user_id:
        can_view = True
    
    if not can_view:
        flash('У вас нет доступа к этому распоряжению', 'danger')
        return redirect(url_for('orders'))
    
    departments = DepartmentModel.get_all()
    dept_users = UserModel.get_by_department(order.get('assigned_department_id')) if order.get('assigned_department_id') else []
    return render_template_string(ORDER_DETAILS_TEMPLATE, order=order, departments=departments, dept_users=dept_users)

@app_flask.route('/orders/<order_id>/status', methods=['POST'])
@login_required
def update_order_status(order_id):
    order = OrderModel.get_by_id(order_id)
    if not order:
        flash('Распоряжение не найдено', 'danger')
        return redirect(url_for('orders'))
    
    new_status = request.form.get('status')
    comment = request.form.get('comment', '')
    user = UserModel.get_by_id(session['user_id'])
    current_role = session['user_role']
    current_status = order['status']
    
    allowed = False
    extra = {}
    details = f'Статус: {new_status}'
    
    if current_role == 'head_central' and current_status == 'На утверждении':
        if new_status in ['Утверждено', 'Отклонено']:
            allowed = True
            if new_status == 'Отклонено' and comment:
                extra['rejection_reason'] = comment
                details = f'Отклонено. Причина: {comment}'
    
    elif current_role == 'secretary' and current_status == 'Утверждено' and new_status == 'В отделе':
        dept_id = request.form.get('department_id')
        if dept_id:
            allowed = True
            extra['assigned_department_id'] = dept_id
            dept = DepartmentModel.get_by_id(dept_id)
            details = f'Назначен отдел: {dept["name"] if dept else dept_id}'
            if dept and dept['head_id']:
                NotificationModel.create(dept['head_id'], f'Распоряжение "{order["title"]}" назначено на ваш отдел', order_id)
    
    elif current_role == 'head_department' and current_status == 'В отделе' and new_status == 'Назначен исполнитель':
        exec_id = request.form.get('executor_id')
        if exec_id:
            allowed = True
            extra['assigned_executor_id'] = exec_id
            executor = UserModel.get_by_id(exec_id)
            details = f'Назначен исполнитель: {executor["full_name"] if executor else exec_id}'
            NotificationModel.create(exec_id, f'Вам назначено распоряжение "{order["title"]}"', order_id)
    
    elif current_role == 'executor' and current_status == 'Назначен исполнитель' and new_status == 'В работе':
        if order.get('assigned_executor_id') == session['user_id']:
            allowed = True
            details = 'Исполнитель взял в работу'
    
    elif current_role == 'head_department' and current_status == 'Готово к проверке' and new_status == 'Подтверждено':
        allowed = True
        details = 'Работа подтверждена начальником отдела'
        head_central_list = UserModel.get_by_role('head_central')
        for hc in head_central_list:
            NotificationModel.create(hc['uid'], f'Распоряжение "{order["title"]}" выполнено, требуется закрытие', order_id)
    
    elif current_role == 'head_central' and current_status == 'Подтверждено' and new_status == 'Закрыто':
        allowed = True
        details = 'Распоряжение закрыто руководителем ЦА'
    
    elif current_role in ['head_department', 'head_central'] and current_status in ['Готово к проверке', 'Подтверждено'] and new_status == 'На доработке':
        allowed = True
        details = f'Отправлено на доработку: {comment}' if comment else 'Отправлено на доработку'
        if order.get('assigned_executor_id'):
            NotificationModel.create(order['assigned_executor_id'], f'Распоряжение "{order["title"]}" отправлено на доработку', order_id)
    
    if allowed:
        OrderModel.update(order_id, status=new_status, **extra)
        OrderHistoryModel.add(order_id, 'Изменение статуса', user['full_name'], current_role, details)
        flash('Статус обновлен', 'success')
    else:
        flash('Действие не разрешено', 'danger')
    
    return redirect(url_for('order_details', order_id=order_id))

@app_flask.route('/orders/<order_id>/submit', methods=['POST'])
@login_required
def submit_order_result(order_id):
    order = OrderModel.get_by_id(order_id)
    if not order:
        flash('Распоряжение не найдено', 'danger')
        return redirect(url_for('orders'))
    
    current_role = session['user_role']
    current_status = order['status']
    result_content = request.form.get('result_content', '').strip()
    
    if current_role != 'executor' or current_status != 'В работе' or order.get('assigned_executor_id') != session['user_id']:
        flash('Действие не разрешено', 'danger')
        return redirect(url_for('order_details', order_id=order_id))
    
    if not result_content:
        flash('Опишите результат выполнения', 'warning')
        return redirect(url_for('order_details', order_id=order_id))
    
    result_file = None
    if 'result_file' in request.files:
        file = request.files['result_file']
        if file and file.filename and allowed_file(file.filename):
            result_file = save_uploaded_file(file, order_id, session['user_id'])
    
    result = {'content': result_content, 'submittedAt': datetime.now().isoformat()}
    OrderModel.update(order_id, status='Готово к проверке', result=result, result_file=result_file)
    
    user = UserModel.get_by_id(session['user_id'])
    OrderHistoryModel.add(order_id, 'Сдача работы', user['full_name'], current_role, 'Работа сдана на проверку')
    
    if order.get('assigned_department_id'):
        dept = DepartmentModel.get_by_id(order['assigned_department_id'])
        if dept and dept['head_id']:
            NotificationModel.create(dept['head_id'], f'Распоряжение "{order["title"]}" сдано на проверку', order_id)
    
    flash('Работа сдана на проверку', 'success')
    return redirect(url_for('order_details', order_id=order_id))

@app_flask.route('/department')
@login_required
def department():
    if session.get('user_role') == 'admin':
        return redirect(url_for('admin_panel'))
    
    dept_id = session.get('department_id')
    if not dept_id:
        flash('У вас нет назначенного отдела', 'warning')
        return redirect(url_for('dashboard'))
    
    department = DepartmentModel.get_by_id(dept_id)
    if not department:
        flash('Отдел не найден', 'danger')
        return redirect(url_for('dashboard'))
    
    users = UserModel.get_by_department(dept_id)
    orders = OrderModel.get_by_department(dept_id)
    return render_template_string(DEPARTMENT_TEMPLATE, department=department, users=users, orders=orders)

@app_flask.route('/admin')
@login_required
@role_required('admin')
def admin_panel():
    users = UserModel.get_all()
    departments = DepartmentModel.get_all()
    return render_template_string(ADMIN_TEMPLATE, users=users, departments=departments)

@app_flask.route('/admin/stats')
@login_required
@role_required('admin')
def admin_stats():
    user_stats = UserModel.get_stats()
    dept_stats = DepartmentModel.get_stats()
    order_stats = OrderModel.get_global_stats()
    completed_orders = order_stats['by_status'].get('Подтверждено', 0) + order_stats['by_status'].get('Закрыто', 0)
    return render_template_string(STATS_TEMPLATE, user_stats=user_stats, dept_stats=dept_stats, 
                                   order_stats=order_stats, completed_orders=completed_orders)

@app_flask.route('/admin/users/create', methods=['POST'])
@login_required
@role_required('admin')
def admin_create_user():
    uid = 'u-' + str(uuid.uuid4())[:8]
    hashed = generate_password_hash(request.form.get('password'))
    UserModel.create(uid, request.form.get('full_name'), request.form.get('email'), 
                     request.form.get('username'), hashed, request.form.get('role'), 
                     request.form.get('department_id') or None)
    flash('Пользователь создан', 'success')
    return redirect(url_for('admin_panel'))

@app_flask.route('/admin/users/<uid>/delete')
@login_required
@role_required('admin')
def admin_delete_user(uid):
    if uid == session['user_id']:
        flash('Нельзя удалить самого себя', 'danger')
    else:
        UserModel.delete(uid)
        flash('Пользователь удален', 'success')
    return redirect(url_for('admin_panel'))

@app_flask.route('/admin/departments/create', methods=['POST'])
@login_required
@role_required('admin')
def admin_create_department():
    dept_id = 'dept-' + str(uuid.uuid4())[:8]
    DepartmentModel.create(dept_id, request.form.get('name'), request.form.get('description'))
    flash('Отдел создан', 'success')
    return redirect(url_for('admin_panel'))

@app_flask.route('/admin/departments/<dept_id>/delete')
@login_required
@role_required('admin')
def admin_delete_department(dept_id):
    DepartmentModel.delete(dept_id)
    flash('Отдел удален', 'success')
    return redirect(url_for('admin_panel'))

@app_flask.route('/admin/export/orders')
@login_required
@role_required('admin')
def admin_export_orders():
    orders = OrderModel.get_all()
    excel_file = export_orders_to_excel(orders)
    if excel_file:
        return send_file(excel_file, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        as_attachment=True, download_name=f'orders_{datetime.now().strftime("%Y%m%d")}.xlsx')
    flash('Ошибка экспорта', 'danger')
    return redirect(url_for('admin_panel'))

@app_flask.route('/admin/export/users')
@login_required
@role_required('admin')
def admin_export_users():
    users = UserModel.get_all()
    excel_file = export_users_to_excel(users)
    if excel_file:
        return send_file(excel_file, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        as_attachment=True, download_name=f'users_{datetime.now().strftime("%Y%m%d")}.xlsx')
    flash('Ошибка экспорта', 'danger')
    return redirect(url_for('admin_panel'))

@app_flask.route('/uploads/<filename>')
def serve_upload(filename):
    filepath = os.path.join(UPLOAD_FOLDER, filename)
    if os.path.exists(filepath):
        return send_file(filepath, as_attachment=True)
    return 'File not found', 404

# ============================================================
# ЗАПУСК
# ============================================================

with app_flask.app_context():
    init_db()
    seed_database()

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app_flask.run(host='0.0.0.0', port=port, debug=False)